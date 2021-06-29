import openpyxl


inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}


# lai noskaidrotu cik produkti ir failaaa
# range lai katru rindu liktu savā ciklā priekš loop
# lai sāktu no 2 rindas liek (2,<- starts -> end (product_list.max_row) <- exclusive neieskaitot pēdējo tapēc + 1
for product_row in range(2, product_list.max_row + 1):
    # product_row, lai katrā loop skatās tekošo rindu un 4 ir statisks jo vajag piegādātāju
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    # Pievienot inventorija  kopējo cenu kolonu
    inventory_price = product_list.cell(product_row, 5)

    # Šeit vajag key un vērtību takā sāks ar 1 tā ir vērtība
    # un šis pievieno biblotekai ar katru inkrementu ja tas nepiecišams
    # aprēķināšana produktu skaitam katram piegādātājam
    if supplier_name in products_per_supplier:
        current_num_product = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_product + 1
    else:
        print("pievieno piegādātāju")
        products_per_supplier[supplier_name] = 1

    # aprēķins kopējai vērtībai visai inventoryam katram piegādātājam:
    # If un els ir lai pievienotu vērtības netikai jaunajiem ierakstiem supplier_name bet arī esošajiem
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # loģika produktiem kur inventorija ir zem 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # Loģika lai pievienotu inventorijas  kopējo cenu

    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)


inv_file.save("inventory_with_total_value.xlsx")
